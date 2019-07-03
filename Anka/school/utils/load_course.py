from openpyxl import load_workbook

USER_INDEXES = ["rut_without_digit", "rut_digit",
                "first_name", "paternal_name", "maternal_name", "role"]

STUDENT_INDEXES = ["rut_without_digit", "rut_digit",
                   "first_name", "paternal_name", "maternal_name"]


def load_course(excel_path):
    wb = load_workbook(excel_path)
    users = wb['Usuarios']
    course_users = []
    for row in users.iter_rows(min_row=2):
        user = {}
        for index, cell in enumerate(row):
            user[USER_INDEXES[index]] = str(cell.value)
        course_users.append(user)

    course = wb['Curso']
    course_info = {'name': course['A2'].value,
                   "institution": course['B2'].value}

    students = wb['Alumnos']
    course_students = []
    for row in students.iter_rows(min_row=2):
        student = {}
        for index, cell in enumerate(row):
            student[STUDENT_INDEXES[index]] = str(cell.value)
        course_students.append(student)

    return course_users, course_info, course_students


if __name__ == "__main__":
    print(load_course('Libro1.xlsx'))
