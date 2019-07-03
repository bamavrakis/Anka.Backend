from django.contrib.auth import get_user_model, authenticate
import numpy as np
from django.utils import timezone
import io
from PIL import Image
import cv2
from .utils.detect_persons import detect_persons
from base64 import b64decode, b64encode
from openpyxl import load_workbook, Workbook
from django.contrib.auth.models import User, Group
from rest_framework import viewsets, response, status
from .models import Teacher, Course, CourseWorker, Student, Attendance
from .serializers import (UserSerializer, AttendanceSerializer,
                          GroupSerializer, TeacherSerializer, CourseSerializer,
                          StudentSerializer, CourseWorkerSerializer)
import datetime
import os
from django.core.mail import send_mail, EmailMessage


User = get_user_model()


class TeacherViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows users to be viewed or edited.
    """
    queryset = Teacher.objects.all()
    serializer_class = TeacherSerializer


class CourseViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows users to be viewed or edited.
    """
    queryset = Course.objects.all()
    serializer_class = CourseSerializer


class StudentViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows users to be viewed or edited.
    """
    queryset = Student.objects.all()
    serializer_class = StudentSerializer


class CourseWorkerViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows users to be viewed or edited.
    """
    queryset = CourseWorker.objects.all()
    serializer_class = CourseWorkerSerializer


class AttendanceViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows users to be viewed or edited.
    """
    queryset = Attendance.objects.all()
    serializer_class = AttendanceSerializer


class GroupViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows groups to be viewed or edited.
    """
    queryset = Group.objects.all()
    serializer_class = GroupSerializer


class LogInViewSet(viewsets.ModelViewSet):
    def create(self, request):
        username = request.data["username"]
        password = request.data["password"]
        try:
            user = Teacher.objects.get(username=username)
            if user.password == password:
                return response.Response(user.id, status=status.HTTP_200_OK)
            return response.Response(status=status.HTTP_400_BAD_REQUEST)
        except Exception:
            return response.Response(status=status.HTTP_400_BAD_REQUEST)


class CheckAttendanceViewSet(viewsets.ModelViewSet):
    def create(self, request):
        image = request.data["base"]
        course_id = request.data["courseId"]
        image_data = b64decode(image)
        data = Image.open((io.BytesIO(image_data)))
        rgb_image = cv2.cvtColor(np.array(data), cv2.COLOR_BGR2RGB)

        course = Course.objects.get(pk=course_id)
        all_students = Student.objects.all()

        detections = detect_persons(rgb_image)
        if not detections:
            return response.Response(status=status.HTTP_400_BAD_REQUEST)

        for student in all_students:
            attended = str(student.rut_without_digit) in detections
            if attended:
                attendance = Attendance(
                    course=course, student=student,
                    status=attended)
                attendance.save()

        return response.Response(status.HTTP_201_CREATED)


class ExportAttendanceViewSet(viewsets.ModelViewSet):
    def create(self, request):
        print(request.data)
        course_id = request.data["courseId"]
        print(course_id)
        wb = Workbook()
        ws = wb.active
        # Data can be assigned directly to cells
        ws['A1'] = "Fecha"
        ws['B1'] = datetime.date.today()
        ws['A2'] = "Rut"
        ws['B2'] = "Nombre"
        ws['C2'] = "Apellido paterno"
        ws['D2'] = "Apellido Materno"
        ws['E2'] = "Asiste"
        course = Course.objects.get(pk=course_id)
        all_students = Student.objects.filter(course=course)
        all_attendances = Attendance.objects.filter(course=course)

        current_index = 3
        for student in all_students:
            ws[f"A{current_index}"] = student.rut_without_digit
            ws[f"B{current_index}"] = student.first_name
            ws[f"C{current_index}"] = student.paternal_name
            ws[f"D{current_index}"] = student.maternal_name
            ws[f"E{current_index}"] = "No"
            for attendance in all_attendances:
                if attendance.date == datetime.date.today() \
                    and attendance.student.rut_without_digit \
                        == student.rut_without_digit:
                    ws[f"E{current_index}"] = "Si"
                    break
            current_index += 1
        wb.save(
            f"school/utils/temp/asistencia_{course.name.strip()}_{datetime.date.today()}.xlsx")
        # ENVIAR MAIL Y LUEGO BORRAR
        email = EmailMessage(
            'Asistencia ' + course.name + " " + str(datetime.date.today()),
            'Estimado docente, \nAdjuntamos la asistencia del día de hoy para el curso ' + course.name + '. \nSaludos.', 'anka@kimche.ai', ['bamavrakis@uc.cl'])
        email.attach_file(
            'school/utils/temp/asistencia_' + course.name.strip() + '_' + str(datetime.date.today()) + '.xlsx')
        email.send()
        os.remove(
            f"school/utils/temp/asistencia_{course.name.strip()}_{datetime.date.today()}.xlsx")
        return response.Response(status.HTTP_201_CREATED)


class CreateEntitiesViewSet(viewsets.ViewSet):
    def create(self, request):
        EXCEL_PATH = "school/utils/Libro1.xlsx"
        USER_INDEXES = ["rut_without_digit", "rut_digit",
                        "first_name", "pate)rnal_name", "maternal_name", "role"]

        STUDENT_INDEXES = ["rut_without_digit", "rut_digit",
                           "first_name", "paternal_name", "maternal_name"]

        wb = load_workbook(EXCEL_PATH)
        users = wb['Usuarios']
        course_users = []
        for row in users.iter_rows(min_row=2):
            user = {}
            for index, cell in enumerate(row):
                user[USER_INDEXES[index]] = str(cell.value)
            course_users.append(user)

        course = wb['Curso']
        course_info = {'name': course['A2'].value,
                       "institution": course['B2'].value}

        students = wb['Alumnos']
        course_students = []
        for row in students.iter_rows(min_row=2):
            student = {}
            for index, cell in enumerate(row):
                student[STUDENT_INDEXES[index]] = str(cell.value)
            course_students.append(student)

        course = Course(**course_info)
        course.save()
        for educator in course_users:
            educator["username"] = educator["rut_without_digit"]
            educator["password"] = educator["rut_without_digit"][0:4]
            user = Teacher(**educator)
            user.save()
            course_worker = CourseWorker(**{"teacher": user, "course": course})
            course_worker.save()

        for student in course_students:
            student["course"] = course
            student = Student(**student)
            student.save()

        return response.Response(status.HTTP_201_CREATED)
